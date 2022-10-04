from openpyxl import load_workbook
from wcmatch import fnmatch

def read_excel(asset_file):
    wb = load_workbook(asset_file)
    dicc={}
    diccionario={}
    ws=wb['Sist. Información - EXT']
    tipos=ws['E1':'E4191']
    paises=ws['C1':'C4191']
    entidades=ws['D1':'D4191']
    activos=ws['F1':'F4191']

    for i in tipos:
        if i[0].value=='App':
            dicc['pais']=paises[i[0].row-1][0].value
            dicc['entidad']=entidades[i[0].row-1][0].value
            dicc['activo']=activos[i[0].row-1][0].value
            diccionario[dicc['activo']]=dicc
            dicc={}
    
    return diccionario