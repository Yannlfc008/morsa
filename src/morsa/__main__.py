from morsa.bing_searcher import BingSearcher
import sys
from openpyxl import load_workbook
from openpyxl import Workbook
from morsa.excel_dominio import read_excel
from morsa.hash import getmd5file
from morsa import NEW_EXCEL_FILE
from morsa import EXCEL_FILE
from morsa import DORK_FILE_PATH
from morsa import CREDENTIALS_FILE_PATH
import logging
from morsa.my_yaml import load_yaml
import click
import requests
from pathlib import Path
import tempfile
import mariadb
import time


@click.command()
@click.option('--output_file',default=NEW_EXCEL_FILE, help='Archivo donde mostrar los resultados')
@click.option('--assets_file',default=EXCEL_FILE,help='Archivo excel con los activos')
@click.option('--dorks_file',default=DORK_FILE_PATH,help='Archivo YML con los dorks')
@click.option('--num_pages',default=None, type=int, help='Límite de páginas donde buscar')
@click.option('--domain',default=[],multiple=True,help='Dominio por el que filtrar al coger los activos')
@click.option('--debug/--no-debug', default=False)
def main(output_file, assets_file, dorks_file,num_pages, domain, debug):
    credentials = load_yaml(CREDENTIALS_FILE_PATH)
    try:
        conn = mariadb.connect(
            user=credentials['CREDENTIALS']['mariadb']['user'],
            password=credentials['CREDENTIALS']['mariadb']['password'],
            host=credentials['CREDENTIALS']['mariadb']['host'],
            port=credentials['CREDENTIALS']['mariadb']['port'],
            database=credentials['CREDENTIALS']['mariadb']['database'])
    except mariadb.Error as e:
        print(f"Error connecting to MariaDB Platform: {e}")
        sys.exit(1)
    
    cur = conn.cursor()
    
    logger = logging.getLogger("morsa")
    logging.basicConfig(filename='morsa_debug.log', filemode='w', format='%(asctime)s|%(levelname)s|%(name)s|%(message)s')
    handler  = logging.StreamHandler()
    handler.setFormatter(logging.Formatter("%(asctime)s|%(levelname)s|%(name)s|%(message)s"))
    logger.addHandler(handler)
    logger.setLevel(logging.DEBUG)
    aux_domain={}
    perros_links=[]
    headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/100.0.4896.60 Safari/537.36'}
    response=None
    count_number_requests=0
    excel_data=read_excel(assets_file)
    file_dorks = load_yaml(dorks_file)
    my_searcher=BingSearcher()
    wb=Workbook()
    #wb.save(NEW_EXCEL_FILE)
    #wb2=load_workbook(NEW_EXCEL_FILE)
    ws=wb.active
    logger.debug("Excel to store results created")
    ws['A1'].value='País'
    ws['B1'].value='Entidad'
    ws['C1'].value='Dominio'
    ws['D1'].value='id_dork'
    ws['E1'].value='riesgo_dork'
    ws['F1'].value='url_resultado'
    ws['G1'].value='hash'
    if debug:
        ws['H1'].value='search'
    wb.save(output_file)
    i=2
    if domain:
        for dominios in domain:
            if dominios in excel_data:
                aux_domain[dominios]=excel_data[dominios]
            else:
                logger.warning("No hay ningun activo con el dominio indicado %s", dominios)
    else:
        aux_domain=excel_data
    
    try:
        for dominio in aux_domain.values():
            for yml_dork in file_dorks['BING']:
                count_number_requests+=1
                dork=yml_dork['dorks']['dork']
                #id_dork=yml_dork['dorks']['id_dork']
                riesgo_dork=yml_dork['dorks']['riesgo']
                country=dominio['pais']
                entidad=dominio['entidad']
                limit=num_pages
                aux_links = my_searcher.search(dominio['activo'],dork,limit)
                #if count_number_requests>1000:
                    #perros_links = my_searcher.search('google.com','perros',1)
                    #if len(perros_links)==0:
                        #raise Exception(count_number_requests)
                if len(aux_links)>0:
                    for url in aux_links:
                        if '.pdf' in url:
                            response=requests.get(url, headers=headers)
                            with tempfile.NamedTemporaryFile() as pdf:
                                pdf.write(response.content)
                                hash_value=getmd5file(pdf.name)
                            try:
                                cur.execute(
                                    "INSERT INTO results (time_stamp,domain,country,entity,dork,risk,uri,md5hash) VALUES (?,?,?,?,?,?,?,?)", 
                                    (time.gmtime(),dominio['activo'],country,entidad,dork,riesgo_dork,url,hash_value))
                                conn.commit()
                            except mariadb.Error as e:
                                print(f"Error: {e}")
                                sys.exit(1)
                            ws['G'+str(i)].value=hash_value
                        else:
                            try:
                                cur.execute(
                                    "INSERT INTO results (time_stamp,domain,country,entity,dork,risk,uri) VALUES (?,?,?,?,?,?,?)", 
                                    (time.gmtime(),dominio['activo'],country,entidad,dork,riesgo_dork,url))
                                conn.commit()
                            except mariadb.Error as e:
                                print(f"Error: {e}")
                                sys.exit(1)
                        ws['A'+str(i)].value=country
                        ws['B'+str(i)].value=entidad
                        ws['C'+str(i)].value=dominio['activo']
                        ws['D'+str(i)].value=dork
                        ws['E'+str(i)].value=riesgo_dork
                        ws['F'+str(i)].value=url
                        if debug:
                            ws['H'+str(i)].value=my_searcher._get_search_query(dominio['activo'],dork)
                        logger.debug("New results added to the excel")
                        i+=1
                wb.save(output_file)
        wb.close()
        conn.close()    
    except KeyboardInterrupt:
        logger.warning("Cogido control+c. Salida segura.")
        wb.close()
    

           


if __name__ == '__main__':
    try:
        main()
    except KeyboardInterrupt:
        print('Exiting..')
        sys.exit()
